import os
from openpyxl import load_workbook
from Bio import SeqIO

tRNA_data = "tRNA data and scoring"
central_directory = "C:\\Users\\sanjay sonney\\SickKids\\Neal Sondheimer - Sanjay Project 2017\\MitoTIP revision work\\MitoTIP"
dataset = "\\dataset files"
output_dir = "\\Output"
os.chdir(central_directory+dataset)
list_tRNA = ["TRNF", "TRNV", "TRNL1", "TRNI", "TRNQ", "TRNM", "TRNW", "TRNA", "TRNN", "TRNC", "TRNY", "TRNS1", "TRND", "TRNK", "TRNG", "TRNR", "TRNH", "TRNS2", "TRNL2", "TRNE", "TRNT", "TRNP"]
dict_mut = {'A': 0, 'G': 1, 'C': 2, 'T': 3}
dict_tRNA_start = {"TRNF": [577, 1, 647],
                        "TRNV": [1602, 1, 1670],
                        "TRNI": [4263, 1, 4331],
                        "TRNM": [4402, 1, 4469],
                        "TRNW": [5512, 1, 5579],
                        "TRND": [7518, 1, 7585],
                        "TRNK": [8295, 1, 8364],
                        "TRNG": [9991, 1, 10058],
                        "TRNR": [10405, 1, 10469],
                        "TRNH": [12138, 1, 12206],
                        "TRNS2": [12207, 1, 12265],
                        "TRNT": [15888, 1, 15953],
                        "TRNP": [16023, 0, 15956],
                        "TRNE": [14742, 0, 14674],
                        "TRNY": [5891, 0, 5826],
                        "TRNC": [5826, 0, 5761],
                        "TRNN": [5729, 0, 5657],
                        "TRNA": [5655, 0, 5587],
                        "TRNL1": [3230, 1, 3304],
                        "TRNL2": [12266, 1, 12336],
                        "TRNS1": [7514, 0, 7446],
                        "TRNQ": [4400, 0, 4329]}
locus_dict = {}
for tRNA in list_tRNA:
    locus_dict[tRNA] = "MT-T" + tRNA[3:]


def find_path_start(start):
    '''
    In the pathogenic mutations reference, this function finds the starting line of the sequences of interest
    '''
    sheet_path_reference = wb['Pathogenic Mutation Reference']
    for rowr in sheet_path_reference.iter_rows('A2:A345'):              #Range needs to be changed depending on if new additions have been made to the file 
        for cellr in rowr:
            if float(cellr.value) >= start: 
                return str(cellr.row)
                
def find_path_end(end):
    '''
    In the pathogenic mutations reference, this function finds the end line of the sequence range of interest
    '''
    sheet_path_reference = wb['Pathogenic Mutation Reference']
    for rowr in sheet_path_reference.iter_rows('A2:A345'):              #Range needs to be changed depending on if new additions have been made to the file 
        for cellr in rowr:
            if cellr.value > end and cellr.value < 16033:               #Highest number must be updated if reference sheet is changed
                return str((cellr.row -1 ))
            elif cellr.value == 16033:                                  #Highest number must be updated if reference sheet is changed
                return str(cellr.row)
                
def find_poly_start(start):
    '''
    In the pathogenic mutations reference, this function finds the starting line of the sequences of interest
    '''
    sheet_poly_reference = wb['Mitomap Polymorphism Reference']
    for rowr in sheet_poly_reference.iter_rows('A2:A9846'):           #Range needs to be changed depending on if new additions have been made to the file 
        for cellr in rowr:
            if float(cellr.value) >= start: 
                # print ("start:" + str(start) + "-" + str(cellr.row))  #diagnostic 
                return str(cellr.row)
                
def find_poly_end(end):
    '''
    In the pathogenic mutations reference, this function finds the end line of the sequence range of interest
    '''
    sheet_poly_reference = wb['Mitomap Polymorphism Reference']
    for rowr in sheet_poly_reference.iter_rows('A2:A9846'):             #Range needs to be changed depending on if new additions have been made to the file 
        for cellr in rowr:
            if cellr.value > end and cellr.value < 16023:               #Highest number must be updated if reference sheet is changed
                return str((cellr.row -1 ))
                # print ("end:" + str(end) + "-" + str(cellr.row))  #diagnostic
            elif cellr.value == 16023:                                  #Highest number must be updated if reference sheet is changed
                # print ("end:" + str(end) + "-" + str(cellr.row))  #diagnostic
                return str(cellr.row)
                


##### loading sequences
dict_seq_tRNA = {}
active_seq_dict = {}
file = open("sequence_file.txt", 'r')           #This is to store the sequence 
list_sequence_file = file.readlines()
file.close()
for i in range(int(len(list_sequence_file)/2)):
    tRNA=i*2
    seq = tRNA+1
    active_seq_dict[list_sequence_file[tRNA].strip("\n")] = list_sequence_file[seq].strip("\n")
for tRNA in list_tRNA:
    dict_seq_tRNA[tRNA] = {}                    # setting up dictionary where with tRNA and sequence you can get the base 
    if dict_tRNA_start[tRNA][1] == 1:           # determining whether to count up or down depending on light or heavy strand   
        Active_seq= active_seq_dict[tRNA]
        for i in range(len(Active_seq)):
            dict_seq_tRNA[tRNA][dict_tRNA_start[tRNA][0]+i] = Active_seq[i]
    else:
        Active_seq= active_seq_dict[tRNA]
        for i in range(len(Active_seq)):
            dict_seq_tRNA[tRNA][dict_tRNA_start[tRNA][2]+i] = Active_seq[i]


####        This clears the sheets before inserting from database 
wb = load_workbook(tRNA_data + '.xlsx')
for tRNA in list_tRNA:
    sheet_active = wb[tRNA]
    if dict_tRNA_start[tRNA][1] ==1: 
        start_num = dict_tRNA_start[tRNA][0]
        end_num = dict_tRNA_start[tRNA][2]
    elif dict_tRNA_start[tRNA][1] ==0: 
        start_num = dict_tRNA_start[tRNA][2]
        end_num = dict_tRNA_start[tRNA][0]
    for rowr in sheet_active.iter_rows(min_row = 1, max_row = 1, min_col = 0, max_col = 4*(end_num-start_num)+5):
        for cell in rowr:
            end_letter = str(cell.column)
    for row in sheet_active.iter_rows('B3:' + end_letter + '3'):
        for cell in row:
            sheet_active.cell(row = 7, column = cell.col_idx, value = "")
            sheet_active.cell(row = 7, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 7, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 7, column = cell.col_idx+3, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx+3, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx+3, value = "")
    print (tRNA + " is cleared!")
    

#####       Goes through database and inserts the appropriate mutations
sheet_path_reference = wb['Pathogenic Mutation Reference']
sheet_poly_reference = wb['Mitomap Polymorphism Reference']

for tRNA in list_tRNA:      #Adds pathogenic mutations 
    sheet_active = wb[tRNA]
    if dict_tRNA_start[tRNA][1] ==1: 
        start_num = dict_tRNA_start[tRNA][0]
        end_num = dict_tRNA_start[tRNA][2]
    elif dict_tRNA_start[tRNA][1] ==0: 
        start_num = dict_tRNA_start[tRNA][2]
        end_num = dict_tRNA_start[tRNA][0]
    search_range = "A" + find_path_start(start_num) + ":A" + find_path_end(end_num)
    for rowr in sheet_active.iter_rows(min_row = 1, max_row = 1, min_col = 0, max_col = 4*(end_num-start_num)+5):
        for cell in rowr:
            end_letter = str(cell.column)
    for rowr in sheet_path_reference.iter_rows(search_range):
        for cellr in rowr:
            if (end_num + 1)> cellr.value > (start_num -1):
                for row in sheet_active.iter_rows('A3:' + end_letter + '3'):
                    for cell in row:
                        if cell.value == cellr.value:
                            ref_resulting_change = str(sheet_path_reference.cell(row = cellr.row, column = 4).value)[1+len(str(cellr.value))]
                            MitoMAP_path_A = sheet_active.cell(row = 7, column = cell.col_idx)
                            MitoMAP_path_G = sheet_active.cell(row = 7, column = cell.col_idx+1)
                            MitoMAP_path_C = sheet_active.cell(row = 7, column = cell.col_idx+2)
                            MitoMAP_path_T = sheet_active.cell(row = 7, column = cell.col_idx+3)
                            MitoMAP_path_ref = str(sheet_path_reference.cell(row = cellr.row, column = 4).value)
                            if ref_resulting_change == 'A':
                                if str(MitoMAP_path_A.value) != 'None' and str(MitoMAP_path_A.value) != "":
                                    sheet_active.cell(row = 7, column = cell.col_idx, value = str(MitoMAP_path_A.value)+ ", " + MitoMAP_path_ref)
                                else:
                                    sheet_active.cell(row = 7, column = cell.col_idx, value = MitoMAP_path_ref)
                            if ref_resulting_change == 'G':
                                if str(MitoMAP_path_G.value) != 'None' and str(MitoMAP_path_G.value) != "":
                                    sheet_active.cell(row = 7, column = cell.col_idx+1, value = str(MitoMAP_path_G.value)+ ", " + MitoMAP_path_ref)
                                else:
                                    sheet_active.cell(row = 7, column = cell.col_idx+1, value = MitoMAP_path_ref)
                            if ref_resulting_change == 'C':
                                if str(MitoMAP_path_C.value) != 'None' and str(MitoMAP_path_C.value) != "":
                                    sheet_active.cell(row = 7, column = cell.col_idx+2, value = str(MitoMAP_path_C.value)+ ", " + MitoMAP_path_ref)
                                else:
                                    sheet_active.cell(row = 7, column = cell.col_idx+2, value = MitoMAP_path_ref)
                            if ref_resulting_change == 'T':
                                if str(MitoMAP_path_T.value) != 'None' and str(MitoMAP_path_T.value) != "":
                                    sheet_active.cell(row = 7, column = cell.col_idx+3, value = str(MitoMAP_path_T.value)+ ", " + MitoMAP_path_ref)
                                else:
                                    sheet_active.cell(row = 7, column = cell.col_idx+3, value = MitoMAP_path_ref)
                            if ref_resulting_change == 'd' or ref_resulting_change == ':' or ref_resulting_change == "i":
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'A':
                                    if str(MitoMAP_path_A.value) != 'None' and str(MitoMAP_path_A.value) != "":
                                        sheet_active.cell(row = 7, column = cell.col_idx, value = str(MitoMAP_path_A.value)+ ", " + MitoMAP_path_ref)
                                    else:
                                        sheet_active.cell(row = 7, column = cell.col_idx, value = MitoMAP_path_ref)
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'G':
                                    if str(MitoMAP_path_G.value) != 'None' and str(MitoMAP_path_G.value) != "":
                                        sheet_active.cell(row = 7, column = cell.col_idx+1, value = str(MitoMAP_path_G.value)+ ", " + MitoMAP_path_ref)
                                    else:
                                        sheet_active.cell(row = 7, column = cell.col_idx+1, value = MitoMAP_path_ref)
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'C':
                                    if str(MitoMAP_path_C.value) != 'None' and str(MitoMAP_path_C.value) != "":
                                        sheet_active.cell(row = 7, column = cell.col_idx+2, value = str(MitoMAP_path_C.value)+ ", " + MitoMAP_path_ref)
                                    else:
                                        sheet_active.cell(row = 7, column = cell.col_idx+2, value = MitoMAP_path_ref)
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'T':
                                    if str(MitoMAP_path_T.value) != 'None' and str(MitoMAP_path_T.value) != "":
                                        sheet_active.cell(row = 7, column = cell.col_idx+3, value = str(MitoMAP_path_T.value)+ ", " + MitoMAP_path_ref) 
                                    else:
                                        sheet_active.cell(row = 7, column = cell.col_idx+3, value = MitoMAP_path_ref)
print ("Done inserting Pathogenic mutations")

for tRNA in list_tRNA:      #Adds poly mutations
    sheet_active = wb[tRNA]
    seq_range = str(sheet_active['A2'].value)
    seq_range = seq_range.replace(" ", "")
    start_num = int (seq_range[0:seq_range.index("-")])
    end_num = int(seq_range[seq_range.index("-")+1:len(seq_range)])
    # print("The start and stop positions extracted from the file are " + str(start_num) + ", and " + str(end_num) + ", respectively.")
    Active_seq= active_seq_dict[tRNA]
    search_range = "A" + find_poly_start(start_num) + ":A" + find_poly_end(end_num)
    for rowr in sheet_active.iter_rows(min_row = 1, max_row = 1, min_col = 0, max_col = 4*(end_num-start_num)+5):
        for cell in rowr:
            end_letter = str(cell.column)
    for rowr in sheet_poly_reference.iter_rows(search_range):
        for cellr in rowr:
            if (end_num + 1)> cellr.value > (start_num -1) and sheet_poly_reference.cell(row = cellr.row, column = 2).value == locus_dict[tRNA]:
                for row in sheet_active.iter_rows('A3:' + end_letter + '3'):
                    for cell in row:
                        if cell.value == cellr.value:
                            ref_resulting_change = sheet_poly_reference.cell(row = cellr.row, column = 3)
                            MitoMAP_poly_A = sheet_active.cell(row = 5, column = cell.col_idx).value
                            MitoMAP_poly_A_f = sheet_active.cell(row = 6, column = cell.col_idx).value
                            MitoMAP_poly_G = sheet_active.cell(row = 5, column = cell.col_idx+1).value
                            MitoMAP_poly_G_f = sheet_active.cell(row = 6, column = cell.col_idx+1).value
                            MitoMAP_poly_C = sheet_active.cell(row = 5, column = cell.col_idx+2).value
                            MitoMAP_poly_C_f = sheet_active.cell(row = 6, column = cell.col_idx+2).value
                            MitoMAP_poly_T = sheet_active.cell(row = 5, column = cell.col_idx+3).value
                            MitoMAP_poly_T_f = sheet_active.cell(row = 6, column = cell.col_idx+3).value
                            MitoMAP_poly_ref = sheet_poly_reference.cell(row = cellr.row, column = 9).value
                            if str(ref_resulting_change.value)[2] == 'A':
                                if str(MitoMAP_poly_A) != 'None' and str(MitoMAP_poly_A) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx, value = str(MitoMAP_poly_A)+", "+ str(ref_resulting_change.value))
                                    sheet_active.cell(row = 6, column = cell.col_idx, value = float(MitoMAP_poly_A_f) + float(MitoMAP_poly_ref))
                                    # print (str(ref_resulting_change.value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx, value = str(ref_resulting_change.value))
                                    # print (str(ref_resulting_change.value))   
                                    sheet_active.cell(row = 6, column = cell.col_idx, value = float(MitoMAP_poly_ref))
                            if str(ref_resulting_change.value)[2] == 'G':
                                if str(MitoMAP_poly_G) != 'None' and str(MitoMAP_poly_G) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(MitoMAP_poly_G)+", "+ str(ref_resulting_change.value))
                                    sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(MitoMAP_poly_G_f) + float(MitoMAP_poly_ref))
                                    # print (str(ref_resulting_change.value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(ref_resulting_change.value))
                                    # print (str(ref_resulting_change.value)) 
                                    sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(MitoMAP_poly_ref))
                            if str(ref_resulting_change.value)[2] == 'C':
                                if str(MitoMAP_poly_C) != 'None' and str(MitoMAP_poly_C) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(MitoMAP_poly_C)+ ", "+str(ref_resulting_change.value))
                                    sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(MitoMAP_poly_C_f) + float(MitoMAP_poly_ref))
                                    # print (str(ref_resulting_change.value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(ref_resulting_change.value))
                                    # print (str(ref_resulting_change.value))
                                    sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(MitoMAP_poly_ref))
                            if str(ref_resulting_change.value)[2] == 'T':
                                if str(MitoMAP_poly_T) != 'None' and str(MitoMAP_poly_T) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(MitoMAP_poly_T)+", "+ str(ref_resulting_change.value)) 
                                    sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(MitoMAP_poly_T_f) + float(MitoMAP_poly_ref))
                                    # print (str(ref_resulting_change.value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(ref_resulting_change.value))
                                    # print (str(ref_resulting_change.value))  
                                    sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(MitoMAP_poly_ref))
                            if str(ref_resulting_change.value)[2] == 'd':
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'A':
                                    if str(MitoMAP_poly_A) != 'None' and str(MitoMAP_poly_A) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx, value = str(MitoMAP_poly_A)+", "+str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx, value = float(MitoMAP_poly_A_f) + float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx, value = str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx, value = float(MitoMAP_poly_ref))
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'G':
                                    if str(MitoMAP_poly_G) != 'None' and str(MitoMAP_poly_G) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx+1, value=str(MitoMAP_poly_G)+", "+str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(MitoMAP_poly_G_f) + float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value))          
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'C':
                                    if str(MitoMAP_poly_C) != 'None' and str(MitoMAP_poly_C) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx+2, value=str(MitoMAP_poly_C)+", "+str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(MitoMAP_poly_C_f) + float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value))          
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'T':
                                    if str(MitoMAP_poly_T) != 'None' and str(MitoMAP_poly_T) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx+3, value=str(MitoMAP_poly_T)+", "+str(ref_resulting_change.value)) 
                                        sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(MitoMAP_poly_T_f) + float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(ref_resulting_change.value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(MitoMAP_poly_ref))
                                        # print (str(ref_resulting_change.value)) 
    # print ("Done inserting polymorphisms for sheet: " + tRNA)
print ("Done inserting polymorphisms")
os.chdir(central_directory+output_dir)
wb.save(tRNA_data + '_updated.xlsx')


