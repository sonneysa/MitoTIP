def find_start(start):
    import os
    os.chdir("C:\\Users\\sanjay sonney\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 1 Summer\\Sequence database\\All programs collected together")
    from openpyxl import load_workbook
    wb = load_workbook('tRNA database alignment test.xlsx')
    sheet_reference = wb['Path_mut_ref']
    for rowr in sheet_reference.iter_rows('A2:A306'):
        for cellr in rowr:
            if float(cellr.value) >= start: 
                return str(cellr.row)
                
def find_end(end):
    import os
    os.chdir("C:\\Users\\sanjay sonney\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 1 Summer\\Sequence database\\All programs collected together")
    from openpyxl import load_workbook
    wb = load_workbook('tRNA database alignment test.xlsx')
    sheet_reference = wb['Path_mut_ref']
    for rowr in sheet_reference.iter_rows('A2:A306'):
        for cellr in rowr:
            if cellr.value > end and cellr.value < 16033: 
                return str((cellr.row -1 ))
            elif cellr.value == 16033:
                return str(cellr.row)
                
list_WS = ["TRNF", "TRNV", "TRNL1", "TRNI", "TRNQ", "TRNM", "TRNW", "TRNA", "TRNN", "TRNC", "TRNY", "TRNS1", "TRND", "TRNK", "TRNG", "TRNR", "TRNH", "TRNS2", "TRNL2", "TRNE", "TRNT", "TRNP"]
dict_mut = {'A': 0, 'G': 1, 'C': 2, 'T': 3}
import os
os.chdir("C:\\Users\\sanjay sonney\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 1 Summer\\Sequence database\\All programs collected together")
from openpyxl import load_workbook
from Bio import SeqIO
workbook_name = str(input("Enter the name of the workbook: "))
wb = load_workbook(workbook_name + '.xlsx')
for sheet_name in list_WS:
    sheet_active = wb[sheet_name]
    seq_range = str(sheet_active['A2'].value)
    seq_range = seq_range.replace(" ", "")
    start_num = int (seq_range[0:seq_range.index("-")])
    end_num = int(seq_range[seq_range.index("-")+1:len(seq_range)])
    print("The start and stop positions extracted from the file are " + str(start_num) + ", and " + str(end_num) + ", respectively.")
    search_range = "A" + find_start(start_num) + ":A" + find_end(end_num)
    for seq_record in SeqIO.parse("sequence.fasta", "fasta"):
        seq_ID = seq_record.id
        seq_seq = (repr(seq_record.seq))
        seq_len = (len(seq_record))
    Active_seq= str(seq_record.seq[(start_num-1):(end_num)])
    for rowr in sheet_active.iter_rows('B1',0,4*(len(Active_seq)-1)):
        for cell in rowr:
            end_letter = str(cell.column)
    for row in sheet_active.iter_rows('B3:' + end_letter + '3'):
        for cell in row:
            sheet_active.cell(row = 12, column = cell.col_idx, value = "")
            sheet_active.cell(row = 13, column = cell.col_idx, value = "")
            sheet_active.cell(row = 12, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 13, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 12, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 13, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 12, column = cell.col_idx+3, value = "")
            sheet_active.cell(row = 13, column = cell.col_idx+3, value = "")
    print (sheet_name + " is cleared!")
    
for sheet_name in list_WS:
# sheet_name = str(input("Enter the name of the sheet which you wish to use: "))
    
    sheet_active = wb[sheet_name]
    sheet_reference = wb['Path_mut_ref']
    
    seq_range = str(sheet_active['A2'].value)
    seq_range = seq_range.replace(" ", "")
    start_num = int (seq_range[0:seq_range.index("-")])
    end_num = int(seq_range[seq_range.index("-")+1:len(seq_range)])
    print("The start and stop positions extracted from the file are " + str(start_num) + ", and " + str(end_num) + ", respectively.")
    
    search_range = "A" + find_start(start_num) + ":A" + find_end(end_num)
    
    
    for seq_record in SeqIO.parse("sequence.fasta", "fasta"):
        seq_ID = seq_record.id
        seq_seq = (repr(seq_record.seq))
        seq_len = (len(seq_record))
    Active_seq= str(seq_record.seq[(start_num-1):(end_num)])
    
    for rowr in sheet_active.iter_rows('B1',0,4*(len(Active_seq)-1)):
        for cell in rowr:
            end_letter = str(cell.column)
    for rowr in sheet_reference.iter_rows(search_range):
        for cellr in rowr:
            if (end_num + 1)> cellr.value > (start_num -1):
                for row in sheet_active.iter_rows('A3:' + end_letter + '3'):
                    for cell in row:
                        if cell.value == cellr.value:
                            ref_resulting_change = str(sheet_reference.cell(row = cellr.row, column = 4).value)[1+len(str(cellr.value))]
                            if ref_resulting_change == 'A':
                                if str(sheet_active.cell(row = 12, column = cell.col_idx).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx).value) != "":
                                    sheet_active.cell(row = 12, column = cell.col_idx, value = str(sheet_active.cell(row = 12, column = cell.col_idx).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    sheet_active.cell(row = 13, column = cell.col_idx, value = float(sheet_active.cell(row = 13, column = cell.col_idx).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                else:
                                    sheet_active.cell(row = 12, column = cell.col_idx, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))   
                                    sheet_active.cell(row = 13, column = cell.col_idx, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                            if ref_resulting_change == 'G':
                                if str(sheet_active.cell(row = 12, column = cell.col_idx+1).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx+1).value) != "":
                                    sheet_active.cell(row = 12, column = cell.col_idx+1, value = str(sheet_active.cell(row = 12, column = cell.col_idx+1).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    sheet_active.cell(row = 13, column = cell.col_idx+1, value = float(sheet_active.cell(row = 13, column = cell.col_idx+1).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                else:
                                    sheet_active.cell(row = 12, column = cell.col_idx+1, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value)) 
                                    sheet_active.cell(row = 13, column = cell.col_idx+1, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                            if ref_resulting_change == 'C':
                                if str(sheet_active.cell(row = 12, column = cell.col_idx+2).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx+2).value) != "":
                                    sheet_active.cell(row = 12, column = cell.col_idx+2, value = str(sheet_active.cell(row = 12, column = cell.col_idx+2).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    sheet_active.cell(row = 13, column = cell.col_idx+2, value = float(sheet_active.cell(row = 13, column = cell.col_idx+2).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                else:
                                    sheet_active.cell(row = 12, column = cell.col_idx+2, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    sheet_active.cell(row = 13, column = cell.col_idx+2, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                            if ref_resulting_change == 'T':
                                if str(sheet_active.cell(row = 12, column = cell.col_idx+3).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx+3).value) != "":
                                    sheet_active.cell(row = 12, column = cell.col_idx+3, value = str(sheet_active.cell(row = 12, column = cell.col_idx+3).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value)) 
                                    sheet_active.cell(row = 13, column = cell.col_idx+3, value = float(sheet_active.cell(row = 13, column = cell.col_idx+3).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                else:
                                    sheet_active.cell(row = 12, column = cell.col_idx+3, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 4).value))  
                                    sheet_active.cell(row = 13, column = cell.col_idx+3, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                            if ref_resulting_change == 'd' or ref_resulting_change == ':' or ref_resulting_change == 'i":
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'A':
                                    if str(sheet_active.cell(row = 12, column = cell.col_idx).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx).value) != "":
                                        sheet_active.cell(row = 12, column = cell.col_idx, value = str(sheet_active.cell(row = 12, column = cell.col_idx).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx, value = float(sheet_active.cell(row = 13, column = cell.col_idx).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    else:
                                        sheet_active.cell(row = 12, column = cell.col_idx, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'G':
                                    if str(sheet_active.cell(row = 12, column = cell.col_idx+1).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx+1).value) != "":
                                        sheet_active.cell(row = 12, column = cell.col_idx+1, value = str(sheet_active.cell(row = 12, column = cell.col_idx+1).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx+1, value = float(sheet_active.cell(row = 13, column = cell.col_idx+1).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    else:
                                        sheet_active.cell(row = 12, column = cell.col_idx+1, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx+1, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))          
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'C':
                                    if str(sheet_active.cell(row = 12, column = cell.col_idx+2).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx+2).value) != "":
                                        sheet_active.cell(row = 12, column = cell.col_idx+2, value = str(sheet_active.cell(row = 12, column = cell.col_idx+2).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx+2, value = float(sheet_active.cell(row = 13, column = cell.col_idx+2).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    else:
                                        sheet_active.cell(row = 12, column = cell.col_idx+2, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx+2, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))          
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'T':
                                    if str(sheet_active.cell(row = 12, column = cell.col_idx+3).value) != 'None' and str(sheet_active.cell(row = 12, column = cell.col_idx+3).value) != "":
                                        sheet_active.cell(row = 12, column = cell.col_idx+3, value = str(sheet_active.cell(row = 12, column = cell.col_idx+3).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 4).value)) 
                                        sheet_active.cell(row = 13, column = cell.col_idx+3, value = float(sheet_active.cell(row = 13, column = cell.col_idx+3).value) + float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                    else:
                                        sheet_active.cell(row = 12, column = cell.col_idx+3, value = str(sheet_reference.cell(row = cellr.row, column = 4).value))
                                        sheet_active.cell(row = 13, column = cell.col_idx+3, value = float(sheet_reference.cell(row = cellr.row, column = 11).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 4).value))  
wb.save(workbook_name + '.xlsx')
print ("Done inserting polymorphisms")

