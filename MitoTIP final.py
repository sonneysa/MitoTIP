### setting controlling the outputs of the program
prepare_tRNA_alignment_insertion = 1        #this updates the tRNA data and scoring file
prepare_all_combinations_detailed = 1       #this updates the file containing scores for all possible tRNA mutations with breakdown of scores
prepare_all_combinations = 1                #this updates the file containing scores for all possible tRNA mutations, but only the final predictive score
prepare_recalculation_of_database = 1       #this prepares the recalculated database with the 'take-one-out' approach to assessing sens and spec
threshold = 13.225                          #The starting threshold at which sensitivity and specificity is assessed 
output_sensitivity_specificity = 1          #this determines whether the sensitivity and specificity will be assessed 
first_run = True                            #default set as true 
cons_reference = "Original_conservation"    #This determines which conservation database is used 
# cons_reference = "Primate_conservation"
try:                                        #this determines if the information has been loaded previously 
    mut_dict.get(588, None)
except NameError:
    pass
else:
    first_run = False

### Variables controlling the weight of each of the factors feeding into the specific score; changed during optimization process 
pathpoly = 6                                # weight of pathogneic and polymorphic mutation hx 
cons = 4                                    # weight of conservation of the site between species 

### Variables controlling the weight of the three sources of information feeding into the final prediction score 
feature = 4                                #scaling of featuer score 
specific_score_scal = 11                   #scaling fo specific score
average_score_scal = 10                    #scaling of average score 
base_score_scal = 2                        #scaling of base score (see flowchart)

def get_predictive_score(pos_num, base, recalc):
    '''
    This function takes in the position of the mutation, the base that it is changed to, and whether a recalculation needs to be done
    recalc can be "None", "Path", or "Poly"
    based on these variable, it calculates the predictive score in an output list
    The output is formatted as follows: [output_list_caption, output_list, output_column_list]
        The output_list_caption contains the following: ["Specific score", "Average score", "Feature score", "Predictive score"]
        output_list contains: [specific_score, average_score, feature_score, predictive_score]
        output_column_list: is a list of all the specific scores that went into calculating the average score, with the average at the end of th list
    '''
    tRNA_list = dict_pos_to_tRNA[pos_num]                                                       #Few positions are presen in multiple tRNA 
    predictive_score = []
    new_column_dict = {}
    average_score = []
    specific_score = []
    feature_score = []
    new_column_list = []
    output_column_list = []
    for tRNAi in tRNA_list:
        column_n = tRNA_num_correlation_rev[tRNAi][pos_num][0]
        new_column = {}
        if recalc == 'poly':
            temp_polyf_dict = poly_score_recalculator(pos_num, base, mut_polyf_dict_stored)     #Function that recalculates the poly score 
        for tRNA in list_tRNA:                                                                  #cycles through all of the tRNAs in alignment 
            if tRNA_num_correlation[tRNA][column_n] != None:                                    
                if recalc == 'path':
                    if tRNA_num_correlation[tRNA][column_n] != pos_num:
                        path_score = mut_dict[tRNA_num_correlation[tRNA][column_n]][1]
                    else:
                        path_score = second_highest(mut_dict_path[tRNA_num_correlation[tRNA][column_n]])
                    poly_score = mut_dict[tRNA_num_correlation[tRNA][column_n]][0]
                elif recalc == 'poly':
                    poly_score = temp_polyf_dict[tRNA_num_correlation[tRNA][column_n]]
                    path_score = mut_dict[tRNA_num_correlation[tRNA][column_n]][1]
                elif recalc == None:
                    path_score = mut_dict[tRNA_num_correlation[tRNA][column_n]][1]
                    poly_score = mut_dict[tRNA_num_correlation[tRNA][column_n]][0]
                if path_score == 0 and poly_score == 0:                                         #no patho or poly, then 3+5*scaled cons score
                    assign_value = base_score_scal+(5*(dict_cons_score_transformed[tRNA][tRNA_num_correlation[tRNA][column_n]]/3))
                elif path_score ==0 and poly_score ==1:                                         #no patho and poly = 1, check transition/transverion 
                    if transition_check(transition_dict, tRNA_num_correlation[tRNA][column_n]):
                        assign_value = (5)
                    else:
                        assign_value = (4)
                elif poly_score > 1 and path_score <= 1:                                        #if poly > 1 & path <2 check transition/transversion 
                    if transition_check(transition_dict, tRNA_num_correlation[tRNA][column_n]):
                        assign_value = (5-(poly_score)) #Higher score if transition 
                    else:
                        assign_value = (4-(poly_score)) #lower score if transverion because tolerating more damaging mutation
                else:                                                                           #This is if (poly is <= 1, and path is = 1) OR path >= 2
                    assign_value = (base_score_scal+path_score) #oiginally 3+path_score
                if path_score != 0 or poly_score !=0:                                           #If either score is not 0, then apply the following rule
                    assign_value = ((assign_value/(4+base_score_scal))*pathpoly + (dict_cons_score_transformed[tRNA][tRNA_num_correlation[tRNA][column_n]]/3)*cons)                                                                             #Applies scaling factors for optimization 
                new_column[tRNA_num_correlation[tRNA][column_n]] = (assign_value)
        new_column_list.append(new_column)
        sum_temp = 0
        count_temp = 0
        for tRNA in list_tRNA:
            if tRNA_num_correlation[tRNA][column_n] != None:
                sum_temp += new_column[tRNA_num_correlation[tRNA][column_n]]
                count_temp +=1
        # print(type(specific_score_scal))
        specific_score.append(specific_score_scal*(new_column[pos_num]/10))
        average_score.append(average_score_scal*(sum_temp/count_temp/10))
        if base != None:
            feature_score.append((base_specific_scoring(pos_num, base, tRNAi)/3)*feature)
        else:                                                                                   #This is to keep the list the same length
            feature_score.append(0)
        predictive_score.append(specific_score[-1] + average_score[-1] + feature_score[-1])
        output_column_list.append(tRNAi)
    specific_score = specific_score[predictive_score.index(max(predictive_score))]              #Picks the maximum possiple predictive score 
    average_score = average_score[predictive_score.index(max(predictive_score))]
    feature_score = feature_score[predictive_score.index(max(predictive_score))]
    new_column = new_column_list[predictive_score.index(max(predictive_score))]
    tRNAi = output_column_list[predictive_score.index(max(predictive_score))]
    predictive_score = max(predictive_score)
    output_list_caption = ["Specific score", "Average score", "Feature score", "Predictive score"]
    output_list = [specific_score, average_score, feature_score, predictive_score]
    output_column_list = []
    column_n = tRNA_num_correlation_rev[tRNAi][pos_num][0]
    for tRNA in list_tRNA:
        if tRNA_num_correlation[tRNA][column_n] != None:
            output_column_list.append([str(tRNA_num_correlation[tRNA][column_n]), str(new_column[tRNA_num_correlation[tRNA][column_n]])])
    output_column_list.append("  ")
    output_column_list.append(["Average", sum(new_column.values())/len(new_column.values())])
    output_total = [output_list_caption, output_list, output_column_list]
    return output_total
    
if first_run == True:
    ### importing things 
    import os
    import csv
    import ast
    # central_directory = "C:\\Users\\Sanjay Sonney\\OneDrive for Business\\University of Toronto Medical School\\Year 1 Summer\\Final Master Files - rebuilt\\"
    # central_directory = "C:\\Users\\sanja\\University of Toronto\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 2 summer\\Summer project\\MitoTIP\\Final MitoTIP uptodate"
    central_directory = "C:\\Users\\sanjay sonney\\SickKids\\Neal Sondheimer - Sanjay Project 2017\\MitoTIP revision work\\MitoTIP"
    dataset = "\\dataset files"
    output_dir = "\\Output"
    os.chdir(central_directory + dataset)
    from openpyxl import load_workbook, Workbook
    
    #### opening necessary files 
    # workbook_name = str(input("Enter the name of the workbook: "))
    workbook_name = "tRNA data and scoring"
    wb = load_workbook(workbook_name + '.xlsx', data_only = True)
    
    #### dictionary set up
    tRNA_dict = {}
    dict_mut = {'A': 0, 'G': 1, 'C': 2, 'T': 3}
    dict_mut_rev = {0:'A', 1:'G', 2:'C', 3:'T'}
    mut_dict = {}
    mut_polyf_dict = {}
    mut_path_mod = {}
    transition_dict = {}
    DNA_trans_dict = {"A":"G", "G":"A", "T":"C","C":"T"}
    
    file = open("dict_complement_pos.txt", "r")                                                 #loading dictionaries 
    dict_complement_pos = {}
    for line in file: 
        dict_complement_pos[line.strip("\n")] = ast.literal_eval(file.readline().strip("\n"))
        
    file = open("dict_tRNA_features_stored.txt", "r")
    dict_tRNA_features = {}
    for line in file: 
        dict_tRNA_features[line.strip("\n")] = ast.literal_eval(file.readline().strip("\n"))
    
    list_tRNA = ["TRNF", "TRNV", "TRNI", "TRNM", "TRNW", "TRND", "TRNK", "TRNG", "TRNR", "TRNH", "TRNS2", "TRNT", "TRNP", "TRNE", "TRNY", "TRNC", "TRNN", "TRNA", "TRNL1", "TRNL2", "TRNS1", "TRNQ"]
    dict_tRNA_start = {"TRNF": [577, 1, 647],
                        "TRNV": [1602, 1, 1670],
                        "TRNI": [4263, 1, 4331],
                        "TRNM": [4402, 1, 4469],
                        "TRNW": [5512, 1, 5579],
                        "TRND": [7518, 1, 7585],
                        "TRNK": [8295, 1, 8364],
                        "TRNG": [9991, 1, 10058],
                        "TRNR": [10405, 1, 10469],
                        "TRNH": [12138, 1, 12206],
                        "TRNS2": [12207, 1, 12265],
                        "TRNT": [15888, 1, 15953],
                        "TRNP": [16023, 0, 15956],
                        "TRNE": [14742, 0, 14674],
                        "TRNY": [5891, 0, 5826],
                        "TRNC": [5826, 0, 5761],
                        "TRNN": [5729, 0, 5657],
                        "TRNA": [5655, 0, 5587],
                        "TRNL1": [3230, 1, 3304],
                        "TRNL2": [12266, 1, 12336],
                        "TRNS1": [7514, 0, 7446],
                        "TRNQ": [4400, 0, 4329]}
    
    sheet_alignment1 = wb['Raw Polymorphism scores']
    sheet_alignment2 = wb['Raw Pathogenic scores']
    sheet_alignment3 = wb['Specific scores']
    sheet_poly_ref = wb['Mitomap Polymorphism Reference']
    sheet_path_ref = wb['Pathogenic Mutation Reference']
    
    align_num = 3
    for tRNA in list_tRNA: #sets up the alignment dictionary
        tRNA_dict[tRNA] = align_num
        align_num += 2
    
    tRNA_num_correlation = {}
    tRNA_num_correlation_rev = {}
    for tRNA in list_tRNA:
        tRNA_num_correlation[tRNA] = {}
        tRNA_num_correlation_rev[tRNA] = {}
        for i in range (1,80):
            tRNA_num_correlation[tRNA][i] = sheet_alignment3.cell(row = (tRNA_dict[tRNA]), column = (i+1)).value
            tRNA_num_correlation_rev[tRNA][sheet_alignment3.cell(row = (tRNA_dict[tRNA]), column = (i+1)).value] = [i]
    
    num_to_tRNA = {}
    for tRNA in list_tRNA:
        for key in list(tRNA_num_correlation_rev[tRNA].keys()):
            num_to_tRNA[key] = tRNA
    
    
    
                        
    ##### creating and loading functions
    def percentile_transformer(dictionary):
        ''' This calculates the percentiles for a dictionary and outputs a dictionary that has been transformed by converting all of the values to points between 1-4 depending on what percentile it is'''
        listovalues = []
        percentile_lims = []
        for key in list(dictionary.keys()):
            if sum(dictionary[key]) != 0:
                listovalues.append(sum(dictionary[key]))
        listovalues.sort()
        percentile_lims.append((len(listovalues))/4)
        percentile_lims.append((len(listovalues))/2)
        percentile_lims.append((len(listovalues))*3/4)
        for i in range(len(percentile_lims)):
            if percentile_lims[i] == int(percentile_lims[i]):
                percentile_lims[i] = (listovalues[int(percentile_lims[i])-1] + listovalues[int(percentile_lims[i])])/2
            else:
                percentile_lims[i] = listovalues[int(percentile_lims[i])]
        for key in list(dictionary.keys()):
            if sum(dictionary[key]) != 0:
                if sum(dictionary[key]) <= percentile_lims[0]:
                    dictionary[key] = 1
                elif sum(dictionary[key]) <= percentile_lims[1]:
                    dictionary[key] = 2
                elif sum(dictionary[key]) <= percentile_lims[2]:
                    dictionary[key] = 3
                else:
                    dictionary[key] = 4
            else:
                dictionary[key] = 0
        # print(percentile_lims)
        for i in range(len(percentile_lims)):
            # print(str(percentile_lims[i]) + ": " + str(percentile_lims[i]*30589))                 #Can be used to see what the percentile limits are
            pass
        return dictionary
    
    def transition_check (mut_dict, cur_pos):                                                       #checks if the change is a transition or not 
        for i in range(4):
            # print (mut_dict[cur_pos][i]) 
            if mut_dict[cur_pos][i] ==2:                                                            #Assigned a 2 if it is a transition 
                return False
        return True
        
    def base_specific_scoring(pos_num, base, tRNAi):
        '''
        Enter the position you are interested in and also the base it is changing into then look for the change:
            either complement to non-complement 
            or complement to non-complement with steric hindrance
            or complement to non-complement without steric hinderance
        check before score, then after score and then take the difference 
        after-score minus the before-score 
        complement has lowest score at 0
        non-complement without steric hinderance is 2 
        non-complement with steric hinderance is 3
        deletion is automatically a 3
        multiply the score with a multiplier based on where the mutation is located in the stem'''
        pairing_score_dict = {"A":{"T":0, "C":2, "A":3, "G":3}, "T":{"A":0, "G":2, "C":2, "T":2}, "G":{"C":0, "T":2, "A":3, "G":3}, "C":{"G":0, "A":2, "T":2, "C":2}}
        dict_complement_base = {"A":"T", "T":"A", "G":"C", "C":"G"}
        #calculation of before score 
        tRNA = tRNAi
        z = 0.625                                                                                   #Previously optimized value
        tRNA_list = dict_pos_to_tRNA[pos_num]
        before_score = {}
        after_score = {}
        difference_score = {}
        final_score = {}
        if dict_complement_pos[tRNA][pos_num] != None:
            if dict_tRNA_start[tRNA][1] == 0:                                                               #checking if heavy or light strand
                base = dict_complement_base[base]
                b_base = dict_complement_base[dict_seq_tRNA[tRNA][pos_num]]
                complement = dict_complement_base[dict_seq_tRNA[tRNA][dict_complement_pos[tRNA][pos_num]]]
            else:
                b_base = dict_seq_tRNA[tRNA][pos_num]
                complement = dict_seq_tRNA[tRNA][dict_complement_pos[tRNA][pos_num]]
        
            before_score[tRNA] = pairing_score_dict[b_base][complement]
            
            after_score[tRNA] = pairing_score_dict[base][complement]
            
            difference_score[tRNA] = after_score[tRNA]-before_score[tRNA]
            if b_base == base:                                                                              #If it is a deletion 
                difference_score[tRNA] = 3
            x = dict_tRNA_features[tRNA][pos_num][0]
            n = dict_tRNA_features[tRNA][pos_num][1]
            final_score[tRNA] = ((1-z)*((x-(n/2))**2/(n/2)**2)+z)*difference_score[tRNA]
        else:
            return 0
        return max(final_score.values())
        
    def poly_score_recalculator(pos_num, base, stored_polyf):                                               #used to recalculate quartile poly scores
        '''
        Takes the position, base and stored dictiony of poly scores
        outputs a dictionary of poly scores after the particular base of interest has been removed and reevaluated 
        '''
        temp_poly_dict = {}
        dict_mut = {'A': 0, 'G': 1, 'C': 2, 'T': 3}
        n = dict_mut[base]
        for key in list(stored_polyf.keys()):
            if key != pos_num: 
                temp_poly_dict[key] = stored_polyf[key]
            else:
                temp_poly_dict[key] = [0,0,0,0]
                for i in range(4):
                    if i != n:
                        temp_poly_dict[key][i] = stored_polyf[key][i]
                    else:
                        temp_poly_dict[key][i] = 0
        temp_poly_dict = percentile_transformer(temp_poly_dict)
        return temp_poly_dict
        
    
    
    def second_highest(list_of_four):                                                                       #Only used in recalculation 
        '''
        Takes a list of 4 numbers
        returned the second highest number 
        '''
        max_temp = 0 
        for i in range(4):
            if list_of_four[i] != max(list_of_four):
                if list_of_four[i]>max_temp:
                    max_temp = list_of_four[i]
        return max_temp
    
    def histograph_maker(temp_list):        #Small function for making histograph 
        temp_dict = {}
        for i in range(len(temp_list)):
            temp_list[i] = round(temp_list[i],0)
            temp_dict[temp_list[i]] = 0
        for i in range(len(temp_list)):
            temp_dict[temp_list[i]]+= 1
        return temp_dict
    ##### loading sequences
    dict_seq_tRNA = {}
    active_seq_dict = {}
    file = open("sequence_file.txt", 'r')           #This is to store the sequence 
    list_sequence_file = file.readlines()
    file.close()
    for i in range(int(len(list_sequence_file)/2)):
        tRNA=i*2
        seq = tRNA+1
        active_seq_dict[list_sequence_file[tRNA].strip("\n")] = list_sequence_file[seq].strip("\n")
    for tRNA in list_tRNA:
        dict_seq_tRNA[tRNA] = {}                    # setting up dictionary where with tRNA and sequence you can get the base 
        if dict_tRNA_start[tRNA][1] == 1:           # determining whether to count up or down depending on light or heavy strand   
            Active_seq= active_seq_dict[tRNA]
            for i in range(len(Active_seq)):
                dict_seq_tRNA[tRNA][dict_tRNA_start[tRNA][0]+i] = Active_seq[i]
        else:
            Active_seq= active_seq_dict[tRNA]
            for i in range(len(Active_seq)):
                dict_seq_tRNA[tRNA][dict_tRNA_start[tRNA][2]+i] = Active_seq[i]
    dict_pos_to_tRNA = {}                           # making a dictionary of position to tRNA 
    for tRNA in list_tRNA:  
        for key in list(dict_seq_tRNA[tRNA].keys()):
            dict_seq_tRNA[tRNA][key]
            if dict_pos_to_tRNA.get(key, None) == None: 
                dict_pos_to_tRNA[key] = []
            dict_pos_to_tRNA[key].append(tRNA)

   #### loading excel file which is a conservation score reference
    dict_cons_score = {}
    os.chdir(central_directory+dataset)
    wbc = load_workbook(cons_reference + '.xlsx', data_only = True)
    for tRNA in list_tRNA: 
        dict_cons_score[tRNA] = {}
        ws1 = wbc[tRNA]
        rw = 1
        while ws1.cell(row = rw, column = 1).value != "del":
            rw+=1
        rw +=1
        cl = 1
        while ws1.cell(row = 2, column = cl).value != "Acceptor-stem end":
            cl +=1
            dict_cons_score[tRNA][ws1.cell(row = (1), column = (cl)).value] = ws1.cell(row = (rw), column = (cl)).value
    all_cons_scores = {}
    all_cons_scores_transformed = {}
    tRNA_dict = {}
    for tRNA in list_tRNA: #sets up the alignment dictionary
        for key in list(dict_cons_score[tRNA].keys()):
            all_cons_scores[key] = dict_cons_score[tRNA][key]       #gets all conservation scores in one dictionary so max and min can be determined
    for key in list(all_cons_scores.keys()):                        #transforms the conservation score by the scaling factor 
        all_cons_scores_transformed[key] = 3*(all_cons_scores[key]-min(all_cons_scores.values()))/(max(all_cons_scores.values())-min(all_cons_scores.values()))
    dict_cons_score_transformed = {}
    for tRNA in list_tRNA:                                          #transforms the conservation score by the scaling facotr
        dict_cons_score_transformed[tRNA] = {}
        for key in list(dict_cons_score[tRNA].keys()):
            dict_cons_score_transformed[tRNA][key] = 3 *(dict_cons_score[tRNA][key]-min(all_cons_scores.values()))/(max(all_cons_scores.values())-min(all_cons_scores.values()))
        
    
    
    ##### Setting alignment number so you know where to start writing and how many to shift down for each line
    list_tRNA = ["TRNF", "TRNV", "TRNI", "TRNM", "TRNW", "TRND", "TRNK", "TRNG", "TRNR", "TRNH", "TRNS2", "TRNT", "TRNP", "TRNE", "TRNY", "TRNC", "TRNN", "TRNA", "TRNL1", "TRNL2", "TRNS1", "TRNQ"]
    tRNA_dict = {}
    align_num = 3
    for tRNA in list_tRNA: #sets up the alignment dictionary
        tRNA_dict[tRNA] = align_num
        align_num += 2
    ##### Done alignment dictionary 
    
    ##### Preparing the sheets, collecting polyf, mut dict
    wb = load_workbook(workbook_name + '.xlsx', data_only = True)
    sheet_alignment1 = wb['Raw Polymorphism scores']
    sheet_alignment2 = wb['Raw Pathogenic scores']
    sheet_alignment3 = wb['Specific scores']
    ##### goes through the mutations in the path_ref file and sets the modification values in the dictionary; 
    for rowr in sheet_path_ref.iter_rows('A2:A322'): #Creates dictionary properly 
        for cellr in rowr:
            if sheet_path_ref.cell(row = cellr.row, column = 5).value[0:4] == "tRNA":
                Empty_mut_dataset = [[0,0,0],[0,0,0],[0,0,0],[0,0,0]]                   
                '''4 sublists represent 4 possible bases (A,G,C,T), 
                3 numbers repersent heteroplasmy, confirmation of mutation, and transversion/transition (not currently used)
                '''
                mut_path_mod[cellr.value] = [[0,0,0],[0,0,0],[0,0,0],[0,0,0]]
    cellr_previous = 0
    for rowr in sheet_path_ref.iter_rows('A2:A322'): #Now the modification values will be changed depending upon the features of the mutation, needs to be udpated
        for cellr in rowr:
            if sheet_path_ref.cell(row = cellr.row, column = 5).value[0:4] == "tRNA":
                cur_cell_val = sheet_path_ref.cell(row = cellr.row, column = 4).value
                cur_cell_str = str(cur_cell_val)
                base_chg_indx = 1+len(str(cellr.value))
                if cur_cell_str[base_chg_indx] == "A" or cur_cell_str[base_chg_indx] == "G" or cur_cell_str[base_chg_indx] == "C" or cur_cell_str[base_chg_indx] == "T":
                    dict_mut_key = dict_mut[cur_cell_str[base_chg_indx]]                #Tells you which sublist to use depending on base (AGCT -> 0123)
                else:
                    dict_mut_key = dict_mut[cur_cell_str[0]]
                if sheet_path_ref.cell(row = cellr.row, column = 7).value == "+":           #If it is heteroplasmic, the first number is changed to a 2 
                    mut_path_mod[cellr.value][dict_mut_key][0] = 2
                if sheet_path_ref.cell(row = cellr.row, column = 8).value == "Cfrm":        #If it is confirmed, second number changed to a 1
                    mut_path_mod[cellr.value][dict_mut_key][1] = 1
                if cur_cell_val[0] == DNA_trans_dict.get(cur_cell_val[base_chg_indx], ""):  #If it is a transition, do the following (currently nothing)
                    mut_path_mod[cellr.value][dict_mut_key][2] = 0 #currently sent to 0 so it doesn't affect pathogenic score
                cellr_previous = cellr.value
        
    #### This goes through all the sheets and makes a dictionary of all the mutations and their associated mutation frequencies, it records pathogenic and polymophic mutations separately
    for sheet_name in list_tRNA: 
        sheet_active = wb[sheet_name]
        seq_range = str(sheet_active['A2'].value)
        seq_range = seq_range.replace(" ", "")
        start_num = int (seq_range[0:seq_range.index("-")])
        end_num = int(seq_range[seq_range.index("-")+1:len(seq_range)])
        print("The start and stop positions extracted from the file are " + str(start_num) + ", and " + str(end_num) + ", respectively.")
        cur_pos = ""
        for rowr in sheet_active.iter_rows(min_row = 1, max_row = 1, min_col = 0, max_col = 4*(end_num-start_num)+5): # Captures the range that it needs to search through
            for cell in rowr:
                end_letter = str(cell.column)
        for row in sheet_active.iter_rows('B3:' + end_letter + '3'):
            for cell in row:
                if cell.value != None: #If the cell is not empty, then capture the value and set up a new dictionary at that value. Set the counter to 0 again for capturing the pathogenic variations. 
                    cur_pos = cell.value
                    mut_dict[cur_pos] = [[0,0,0,0],[0,0,0,0]]
                    mut_polyf_dict[cur_pos] = [0,0,0,0]
                    mut_dict_n = 0
                MitoMAP_poly = sheet_active.cell(row = 5, column = cell.col_idx)
                MitoMAP_poly_f = sheet_active.cell(row = 6, column = cell.col_idx)
                MitoMAP_path = sheet_active.cell(row = 7, column = cell.col_idx)
                MAMit_path = sheet_active.cell(row = 8, column = cell.col_idx)
                Pubmed_path = sheet_active.cell(row = 10, column = cell.col_idx)
                if cell.value == cur_pos or cell.value == None: #This adds polymorphic occurance, polymorphic frequency, and pathogenic occurance
                    if MitoMAP_poly_f.value != None: #This checks if polymorphic is recorded
                        if MitoMAP_poly_f.value != 0 and MitoMAP_poly_f.value != None: #record the frequency as follows
                            mut_polyf_dict[cur_pos][mut_dict_n] = MitoMAP_poly_f.value
                        else: #the only possibility is that there is no frequency data from mitomap
                            mut_polyf_dict[cur_pos][mut_dict_n] = 1/30589
                    if MitoMAP_poly.value != None: #If there is a polymorphic mutation is recorded like this, as a plus 1 at that particular position
                        if MitoMAP_poly.value[0] == DNA_trans_dict.get(MitoMAP_poly.value[-1], ""):     #If it is a transition, assign a 1
                            mut_dict[cur_pos][0][mut_dict_n] = 1
                        else:                                                                           #If it is a transversion, assign a 2
                            mut_dict[cur_pos][0][mut_dict_n] = 2
                    if MitoMAP_path.value != None or Pubmed_path.value != None or MAMit_path.value != None: #If there is a pathogenic mutation: 
                        mut_dict[cur_pos][1][mut_dict_n] += 1
                    mut_dict_n += 1
        print ("Done writing dictionary for " + sheet_name) 
        
    
    ######### What this does is set up a dictionary for determining transition vs transversion when calculating poly score 
    for pos_num in sorted(list(mut_dict.keys())):
        transition_dict[pos_num]= mut_dict[pos_num][0]
        
    ######### This tranforms the frequencies to percentiles 
    mut_polyf_dict_stored = {}
    for key in list(mut_polyf_dict.keys()):
        mut_polyf_dict_stored[key] = mut_polyf_dict[key]
        
    mut_polyf_dict = percentile_transformer(mut_polyf_dict)
    
    ######## This modifies the mut_dict by the features of the pathogenic mutation
    mut_dict_path = {}
    for pos_num in sorted(list(mut_dict.keys())): #For each thing in the dictionary, the percentiles are scored and put into the polymorphic position, and the pathogenic positions are scored based on the maximum score obtained for each mutation at each position.
        mut_dict[pos_num][0] = mut_polyf_dict[pos_num]              #Transfered the percentile score to the mut_dict main dictionary
        for i in range(4):
            mut_dict[pos_num][1][i] = mut_dict[pos_num][1][i] + sum(mut_path_mod.get(pos_num, Empty_mut_dataset)[i])
        mut_dict_path[pos_num] = mut_dict[pos_num][1]               #Stores the pathogenic score in a separate dictionary 
        mut_dict[pos_num][1] = max(mut_dict[pos_num][1])            #Selects the maximum mutation score to represent that position
    
    ##### Finished preparing sheets, setting up mut dict


##### For insertion into master database 
if prepare_tRNA_alignment_insertion == 1: 
    for sheet_name in list_tRNA:
        for rowa in sheet_alignment1.iter_rows('B'+ str(tRNA_dict[sheet_name]) + ":CB" + str(tRNA_dict[sheet_name])):
            for cell in rowa:
                if cell.value != None:
                    sheet_alignment1.cell(row = cell.row +1, column = cell.col_idx, value = (mut_dict[cell.value][0])) 
    print ("Done writing alignment for Raw polymoprhism score")
    
    for sheet_name in list_tRNA:
        for rowa in sheet_alignment2.iter_rows('B'+ str(tRNA_dict[sheet_name]) + ":CB" + str(tRNA_dict[sheet_name])):
            for cell in rowa:
                if cell.value != None:
                    sheet_alignment2.cell(row = cell.row +1, column = cell.col_idx, value = (mut_dict[cell.value][1])) #this was modified just to check was pathogenic only would look like
    print ("Done writing alignment for Raw pathogenic score")
    
    for sheet_name in list_tRNA:
        for rowa in sheet_alignment3.iter_rows('B'+ str(tRNA_dict[sheet_name]) + ":CB" + str(tRNA_dict[sheet_name])):
            for cell in rowa:
                if cell.value != None:
                    pos_num = cell.value
                    base = None
                    recalc = None
                    output = get_predictive_score(pos_num, base, recalc)
                    specific_score = output[1][0]
                    sheet_alignment3.cell(row = cell.row +1, column = cell.col_idx, value = specific_score)
    print ("Done writing alignment for Predictive score")
    os.chdir(central_directory + output_dir)
    wb.save(workbook_name + '_scored.xlsx')

##### For the recalculation (take one out approach) and calculation of sensitivity and specificity 
if prepare_recalculation_of_database == 1:
    os.chdir(central_directory + output_dir)
    list_of_mut = []
    for pos_num in list(mut_dict.keys()):
        if mut_dict[pos_num][1]>3:    #only confirmed mutations are being tested now
        # if mut_dict[pos_num][1]>1:      #To test confirmed and/or heteroplasmic mutations 
            list_of_mut.append([pos_num, dict_mut_rev[mut_dict_path[pos_num].index(max(mut_dict_path[pos_num]))]])
    list_of_mut.sort()
    classified_mut = 0
    
    
    list_of_poly = []
    for pos_num in list(mut_dict.keys()):
        if mut_dict[pos_num][0]>0 and mut_dict[pos_num][1]==0:  #positions with polymoprhisms but not pathogenic mutations
            for i in range (4):
                if mut_polyf_dict_stored[pos_num][i] != 0:
                    list_of_poly.append([pos_num, dict_mut_rev[i]])
    list_of_poly.sort()
    classified_poly = 0 
    
    output_predictive = {}
    output_predictive['path'] = {}
    output_predictive['poly'] = {}
    
    wb = Workbook()
    ws1 = wb.create_sheet()
    ws1.title = "Pathogenic mutations"
    
    temp_list = []
    for i in range(len(list_of_mut)):
        pos_num = list_of_mut[i][0]
        base = list_of_mut[i][1]
        recalc = 'path'
        output = get_predictive_score(pos_num, base, recalc)
        ws1.cell(row = (i+2), column = 1, value = pos_num)
        ws1.cell(row = (i+2), column = 2, value = base)
        ws1.cell(row = (i+2), column = 3, value = output[1][0])
        ws1.cell(row = (i+2), column = 4, value = output[1][1])
        ws1.cell(row = (i+2), column = 5, value = output[1][2])
        ws1.cell(row = (i+2), column = 6, value = output[1][3])
        if output_predictive['path'].get(pos_num, None) == None:
            output_predictive['path'][pos_num] = {}
        output_predictive['path'][pos_num][base] = output[1][3]
        if output[1][3]>threshold:
            classified_mut += 1
        temp_list.append(output[1][3])
        
    path_histo_dict = histograph_maker(temp_list)
    for i in range(len(list(path_histo_dict.keys()))):
        ws1.cell(row = (i+2), column = 7, value = list(path_histo_dict.keys())[i])
        ws1.cell(row = (i+2), column = 8, value = path_histo_dict[list(path_histo_dict.keys())[i]])
        
    
    ws1 = wb.create_sheet()
    ws1.title = "Polymorphic mutations"
    
    temp_list = []
    for i in range(len(list_of_poly)):
        pos_num = list_of_poly[i][0]
        base = list_of_poly[i][1]
        recalc = 'poly'
        output = get_predictive_score(pos_num, base, recalc)
        ws1.cell(row = (i+2), column = 1, value = pos_num)
        ws1.cell(row = (i+2), column = 2, value = base)
        ws1.cell(row = (i+2), column = 3, value = output[1][0])
        ws1.cell(row = (i+2), column = 4, value = output[1][1])
        ws1.cell(row = (i+2), column = 5, value = output[1][2])
        ws1.cell(row = (i+2), column = 6, value = output[1][3])
        if output_predictive['poly'].get(pos_num, None) == None:
            output_predictive['poly'][pos_num] = {}
        output_predictive['poly'][pos_num][base] = output[1][3]
        if output[1][3]<=threshold:
            classified_poly += 1
        temp_list.append(output[1][3])
        
    poly_histo_dict = histograph_maker(temp_list)
    for i in range(len(list(poly_histo_dict.keys()))):
        ws1.cell(row = (i+2), column = 7, value = list(poly_histo_dict.keys())[i])
        ws1.cell(row = (i+2), column = 8, value = poly_histo_dict[list(poly_histo_dict.keys())[i]])
    os.chdir(central_directory + output_dir)
    wb.save("Recalculated_scores" + '.xlsx')
    print ("Sensitivity:" + str(classified_mut/len(list_of_mut)))
    print("Specificity: " + str(classified_poly/len(list_of_poly)))
    
    def get_sens_spec (threshold, output_predictive, show_print):       #Function to optimize sensitivity and specificity 
        classified_mut = 0
        is_mut = 0
        classified_poly = 0
        is_poly = 0
        for pos_num in list(output_predictive['path'].keys()): #checking the pathogenics first
            for base in list(output_predictive['path'][pos_num].keys()):
                is_mut += 1
                predictive_score = output_predictive['path'][pos_num][base]
                if predictive_score > threshold:
                    classified_mut += 1
                    
        for pos_num in list(output_predictive['poly'].keys()): #checking the pathogenics first
            for base in list(output_predictive['poly'][pos_num].keys()):
                is_poly += 1
                predictive_score = output_predictive['poly'][pos_num][base]
                if predictive_score <= threshold:
                    classified_poly += 1
        # print("classified mut: " + str(classified_mut))
        # print ("Is mut: " + str(is_mut))
        # print ("Classified poly: " + str(classified_poly))
        # print("Is poly: " + str(is_poly))
        if show_print == True:
            print ("Sensitivity:" + str(classified_mut/is_mut))
            print("Specificity: " + str(classified_poly/is_poly))
            print(((classified_mut/is_mut + classified_poly/is_poly) + abs(classified_mut/is_mut-classified_poly/is_poly))/(classified_mut/is_mut + classified_poly/is_poly))
        return (((classified_mut/is_mut + classified_poly/is_poly) + abs(classified_mut/is_mut-classified_poly/is_poly))/(classified_mut/is_mut + classified_poly/is_poly))
    
##### All combintations Mitomap output 

if prepare_all_combinations == 1:
    os.chdir(central_directory + output_dir)
    recalc = None
    with open(("All_combinations" +".csv"), "w") as csv_file:
        writer = csv.writer(csv_file, lineterminator = '\n')
        writer.writerow(["Position", "rCRS base", "Change", "Predictive score"])
        output_list = []
        for tRNA in list_tRNA:
            for pos_num in list(dict_seq_tRNA[tRNA].keys()):
                if pos_num != 8365 and pos_num != 8366:
                    for base in list(dict_mut.keys()):
                        output = [get_predictive_score(pos_num, base, recalc)[1][3]]
                        if base != dict_seq_tRNA[tRNA][pos_num]:
                            output.insert(0, base)
                        else:
                            output.insert(0, 'del')
                        output.insert(0, dict_seq_tRNA[tRNA][pos_num])
                        output.insert(0, pos_num)
                        output_list.append (output) 
        dict_output = {}
        n=0
        dict_output[output_list[0][0]] = {}
        for i in range(len(output_list)):
            pos_num = output_list[i][0]
            if n == 4: 
                dict_output[pos_num] = {}
                n = 0
            rCS_base = output_list[i][1]
            Change = output_list[i][2]
            score = output_list[i][3]
            dict_output[pos_num][Change] = output_list[i]
            n += 1
        for pos_num in sorted(dict_output.keys()): 
            for base in sorted(dict_output[pos_num].keys()):
                writer.writerow(dict_output[pos_num][base])
        
##### All combinations detailed output
if prepare_all_combinations_detailed == 1:
    os.chdir(central_directory + output_dir)
    recalc = None
    with open(("All_combinations_detailed" +".csv"), "w") as csv_file:
        writer = csv.writer(csv_file, lineterminator = '\n')
        writer.writerow(["Position", "rCRS base", "Change", "Specific position score", "Aligned average score", "Feature score", "Predictive score"])
        output_list = []
        for tRNA in list_tRNA:
            for pos_num in list(dict_seq_tRNA[tRNA].keys()):
                if pos_num != 8365 and pos_num != 8366:
                    for base in list(dict_mut.keys()):
                        temp_output =get_predictive_score(pos_num, base, recalc)[1]
                        output = [temp_output[0], temp_output[1], temp_output[2], temp_output[3]]
                        if base != dict_seq_tRNA[tRNA][pos_num]:
                            output.insert(0, base)
                        else:
                            output.insert(0, 'del')
                        output.insert(0, dict_seq_tRNA[tRNA][pos_num])
                        output.insert(0, pos_num)
                        output_list.append (output) 
        dict_output = {}
        n=0
        dict_output[output_list[0][0]] = {}
        for i in range(len(output_list)):
            pos_num = output_list[i][0]
            if n == 4: 
                dict_output[pos_num] = {}
                n = 0
            rCS_base = output_list[i][1]
            Change = output_list[i][2]
            score = output_list[i][3]
            dict_output[pos_num][Change] = output_list[i]
            n += 1
        for pos_num in sorted(dict_output.keys()): 
            for base in sorted(dict_output[pos_num].keys()):
                writer.writerow(dict_output[pos_num][base])
