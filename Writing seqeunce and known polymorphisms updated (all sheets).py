def find_start(start, workbook_name):
    import os
    os.chdir("C:\\Users\\sanjay sonney\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 1 Summer\\Sequence database\\All programs collected together")
    from openpyxl import load_workbook
    wb = load_workbook(workbook_name + '.xlsx')
    sheet_reference = wb['Mitomap Reference']
    for rowr in sheet_reference.iter_rows('A2:A9229'):
        for cellr in rowr:
            if cellr.value >= start: 
                return str(cellr.row)
                
def find_end(end, workbook_name):
    import os
    os.chdir("C:\\Users\\sanjay sonney\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 1 Summer\\Sequence database\\All programs collected together")
    from openpyxl import load_workbook
    wb = load_workbook(workbook_name + '.xlsx')
    sheet_reference = wb['Mitomap Reference']
    for rowr in sheet_reference.iter_rows('A2:A9229'):
        for cellr in rowr:
            if cellr.value > end and cellr.value < 16023: 
                return str((cellr.row -1 ))
            elif cellr.value == 16023:
                return str(cellr.row)
    

import os
os.chdir("C:\\Users\\sanjay sonney\\OneDrive - University of Toronto\\University of Toronto Medical School\\Year 1 Summer\\Sequence database\\All programs collected together")
from openpyxl import load_workbook
workbook_name = str(input("Enter the name of the workbook: "))
wb = load_workbook(workbook_name + '.xlsx')
sheet_reference = wb['Mitomap Reference']
list_WS = ["TRNF", "TRNV", "TRNL1", "TRNI", "TRNQ", "TRNM", "TRNW", "TRNA", "TRNN", "TRNC", "TRNY", "TRNS1", "TRND", "TRNK", "TRNG", "TRNR", "TRNH", "TRNS2", "TRNL2", "TRNE", "TRNT", "TRNP"]
locus_dict = {}
for sheet_name in list_WS:
    locus_dict[sheet_name] = "MT-T" + sheet_name[3:]
dict_mut = {'A': 0, 'G': 1, 'C': 2, 'T': 3}
wb = load_workbook(workbook_name +'.xlsx')
from Bio import SeqIO
for seq_record in SeqIO.parse("sequence.fasta", "fasta"):
    seq_ID = seq_record.id
    seq_seq = (repr(seq_record.seq))
    seq_len = (len(seq_record))
    
for sheet_name in list_WS:
    sheet_active = wb[sheet_name]
    seq_range = str(sheet_active['A2'].value)
    seq_range = seq_range.replace(" ", "")
    start_num = int (seq_range[0:seq_range.index("-")])
    end_num = int(seq_range[seq_range.index("-")+1:len(seq_range)])
    print("The start and stop positions extracted from the file are " + str(start_num) + ", and " + str(end_num) + ", respectively.")
    search_range = "A" + find_start(start_num, workbook_name) + ":A" + find_end(end_num, workbook_name)
    for seq_record in SeqIO.parse("sequence.fasta", "fasta"):
        seq_ID = seq_record.id
        seq_seq = (repr(seq_record.seq))
        seq_len = (len(seq_record))
    Active_seq= str(seq_record.seq[(start_num-1):(end_num)])
    for rowr in sheet_active.iter_rows('B1',0,4*(len(Active_seq)-1)):
        for cell in rowr:
            end_letter = str(cell.column)
    for row in sheet_active.iter_rows('B3:' + end_letter + '3'):
        for cell in row:
            sheet_active.cell(row = 5, column = cell.col_idx, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx+1, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx+2, value = "")
            sheet_active.cell(row = 5, column = cell.col_idx+3, value = "")
            sheet_active.cell(row = 6, column = cell.col_idx+3, value = "")
    print (sheet_name + " is cleared!")
    
for sheet_name in list_WS:
    sheet_active = wb[sheet_name]
    seq_range = str(sheet_active['A2'].value)
    seq_range = seq_range.replace(" ", "")
    start_num = int (seq_range[0:seq_range.index("-")])
    end_num = int(seq_range[seq_range.index("-")+1:len(seq_range)])
    print("The start and stop positions extracted from the file are " + str(start_num) + ", and " + str(end_num) + ", respectively.")
    Active_seq= str(seq_record.seq[(start_num-1):(end_num)])
    search_range = "A" + find_start(start_num, workbook_name) + ":A" + find_end(end_num, workbook_name)
    for rowr in sheet_active.iter_rows('B1',0,4*(len(Active_seq)-1)):
        for cell in rowr:
            end_letter = str(cell.column)
    for rowr in sheet_reference.iter_rows(search_range):
        for cellr in rowr:
            if (end_num + 1)> cellr.value > (start_num -1) and sheet_reference.cell(row = cellr.row, column = 2).value == locus_dict[sheet_name]:
                for row in sheet_active.iter_rows('A3:' + end_letter + '3'):
                    for cell in row:
                        if cell.value == cellr.value:
                            if str(sheet_reference.cell(row = cellr.row, column = 3).value)[2] == 'A':
                                if str(sheet_active.cell(row = 5, column = cell.col_idx).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx).value) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx, value = str(sheet_active.cell(row = 5, column = cell.col_idx).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    sheet_active.cell(row = 6, column = cell.col_idx, value = float(sheet_active.cell(row = 6, column = cell.col_idx).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))   
                                    sheet_active.cell(row = 6, column = cell.col_idx, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                            if str(sheet_reference.cell(row = cellr.row, column = 3).value)[2] == 'G':
                                if str(sheet_active.cell(row = 5, column = cell.col_idx+1).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx+1).value) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(sheet_active.cell(row = 5, column = cell.col_idx+1).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(sheet_active.cell(row = 6, column = cell.col_idx+1).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value)) 
                                    sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                            if str(sheet_reference.cell(row = cellr.row, column = 3).value)[2] == 'C':
                                if str(sheet_active.cell(row = 5, column = cell.col_idx+2).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx+2).value) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(sheet_active.cell(row = 5, column = cell.col_idx+2).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(sheet_active.cell(row = 6, column = cell.col_idx+2).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                            if str(sheet_reference.cell(row = cellr.row, column = 3).value)[2] == 'T':
                                if str(sheet_active.cell(row = 5, column = cell.col_idx+3).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx+3).value) != "":
                                    sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(sheet_active.cell(row = 5, column = cell.col_idx+3).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value)) 
                                    sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(sheet_active.cell(row = 6, column = cell.col_idx+3).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                else:
                                    sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    print (str(sheet_reference.cell(row = cellr.row, column = 3).value))  
                                    sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                            if str(sheet_reference.cell(row = cellr.row, column = 3).value)[2] == 'd':
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'A':
                                    if str(sheet_active.cell(row = 5, column = cell.col_idx).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx).value) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx, value = str(sheet_active.cell(row = 5, column = cell.col_idx).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx, value = float(sheet_active.cell(row = 6, column = cell.col_idx).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'G':
                                    if str(sheet_active.cell(row = 5, column = cell.col_idx+1).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx+1).value) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(sheet_active.cell(row = 5, column = cell.col_idx+1).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(sheet_active.cell(row = 6, column = cell.col_idx+1).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx+1, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+1, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value))          
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'C':
                                    if str(sheet_active.cell(row = 5, column = cell.col_idx+2).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx+2).value) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(sheet_active.cell(row = 5, column = cell.col_idx+2).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(sheet_active.cell(row = 6, column = cell.col_idx+2).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx+2, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+2, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value))          
                                if str(sheet_active.cell(row = 4, column = cell.col_idx).value) == 'T':
                                    if str(sheet_active.cell(row = 5, column = cell.col_idx+3).value) != 'None' and str(sheet_active.cell(row = 5, column = cell.col_idx+3).value) != "":
                                        sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(sheet_active.cell(row = 5, column = cell.col_idx+3).value)+ ", " + str(sheet_reference.cell(row = cellr.row, column = 3).value)) 
                                        sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(sheet_active.cell(row = 6, column = cell.col_idx+3).value) + float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                    else:
                                        sheet_active.cell(row = 5, column = cell.col_idx+3, value = str(sheet_reference.cell(row = cellr.row, column = 3).value))
                                        sheet_active.cell(row = 6, column = cell.col_idx+3, value = float(sheet_reference.cell(row = cellr.row, column = 9).value))
                                        print (str(sheet_reference.cell(row = cellr.row, column = 3).value)) 
    print ("Done inserting polymorphisms for sheet: " + sheet_name)
wb.save(workbook_name + '.xlsx')
print ("Done inserting polymorphisms")
    